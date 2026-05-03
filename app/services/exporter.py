from __future__ import annotations

import io

from openpyxl import Workbook

from app.services.transform import INTERNAL_HEADERS, InternalRow

def export_internal_xlsx(rows: list[InternalRow]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "内控结算表"

    ws.append(INTERNAL_HEADERS)
    header_to_col = {h: i + 1 for i, h in enumerate(INTERNAL_HEADERS)}  # 1-based
    two_decimal_headers = {
        "结算日单价(不含税)##jsrdjbhs1",
        "应出勤工时(天)##ycqgst",
        "结算工时(天)##jsgst1",
        "工时费(不含税)##gsfbhs1",
        "加班工时(天)##jbgst1",
        "加班费(不含税)##jbfbhs1",
        "差旅费(不含税)##clfbhs1",
        "差旅补贴(不含税)##clbtbhs1",
    }
    two_decimal_cols = sorted({header_to_col[h] for h in two_decimal_headers if h in header_to_col})

    for r in rows:
        ws.append([r.values.get(h, "") for h in INTERNAL_HEADERS])
        row_idx = ws.max_row
        for col_idx in two_decimal_cols:
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value not in ("", None):
                cell.number_format = "0.00"

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

